"""
Excel CRUD operations
"""

import os
from openpyxl import Workbook, load_workbook

from config.config import FILENAME, HEADERS


def initialize_excel():
    if not os.path.exists(FILENAME):
        wb = Workbook()
        ws = wb.active

        ws.title = "Employees"
        ws.append(HEADERS)
        wb.save(FILENAME)

        print("New Excel file created with headers.")
    else:
        print("Excel file already exists.")


def read_all():
    wb = load_workbook(FILENAME)
    ws = wb.active

    employees = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        employees.append(dict(zip(HEADERS, row)))

    return employees


def create_employee(emp_data):
    wb = load_workbook(FILENAME)
    ws = wb.active

    ws.append(emp_data)
    wb.save(FILENAME)

    return {"message": "Created new employee record."}


def update_employee(emp_id, updated_data):
    wb = load_workbook(FILENAME)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == emp_id:
            for col_idx, header in enumerate(HEADERS):
                if header in updated_data:
                    row[col_idx].value = updated_data[header]

            wb.save(FILENAME)
            return {"message": f"Updated Employee {emp_id} with {updated_data}"}

    return {"message": f"Employee {emp_id} not found."}


def delete_employee(emp_id):
    wb = load_workbook(FILENAME)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == emp_id:
            ws.delete_rows(row)
            wb.save(FILENAME)

            return {"message": f"Deleted Employee {emp_id}"}

    return {"message": f"Employee {emp_id} not found."}
